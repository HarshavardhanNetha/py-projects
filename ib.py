from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
from time import sleep
from selenium.webdriver.common.keys import Keys

data = openpyxl.load_workbook("Mahitha_iB.xlsx")
sheet = data.active

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:8080")
#Change chrome driver path accordingly
chrome_driver = "C:\\Users\\anush\\Downloads\\Harsha\\Class\\Py\\Bot\\chromedriver.exe"
driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)
print(driver.title)

for i in range(70,102):
    driver.find_element_by_xpath("//button[@data-testid=\"create-new\"]").click()

    driver.find_element_by_xpath("//input[@label=\"Person Name\"]").send_keys("{}".format(sheet.cell(i,1).value))
    a=sheet.cell(i,2).value
    age=driver.find_element_by_xpath("//input[@label=\"Age\"]")
    age.send_keys("{}".format(str(a)[:-2]))
    if(sheet.cell(i,3).value=="Male"):
        age.send_keys(Keys.TAB,2*Keys.ARROW_DOWN,Keys.ENTER)
    else:
        age.send_keys(Keys.TAB,Keys.ARROW_DOWN,Keys.ENTER)
    
    #sleep(3)
    #driver.find_element_by_xpath("//div[@class=\"styledComponents__SelectField-sc-1es9x1i-5 DropDown___StyledSelectField-sc-1zpqvh-2 hObhAq css-2b097c-container\"]").click()
    #dropdown to be added
    t=sheet.cell(i,4).value
    driver.find_element_by_xpath("//input[@type=\"tel\"]").send_keys("{}".format(str(t)[:-2]))
    city=driver.find_element_by_xpath("//input[@label=\"Current City/Town/Village\"]")
    city.send_keys("{}".format(sheet.cell(i,5).value))
    #sleep(2)
    
    if(sheet.cell(i,6).value=="School student"):
        city.send_keys(Keys.TAB,Keys.ARROW_DOWN,Keys.ENTER)
    elif(sheet.cell(i,6).value=="College student"):
        city.send_keys(Keys.TAB,2*Keys.ARROW_DOWN,Keys.ENTER)
        #city.send_keys("Keys.ARROW_DOWN")
    elif(sheet.cell(i,6).value=="Parent"):
        city.send_keys(Keys.TAB,3*Keys.ARROW_DOWN,Keys.ENTER)
    elif(sheet.cell(i,6).value=="Teacher"):
        par="Teacher"
        city.send_keys(Keys.TAB,Keys.ARROW_UP,Keys.ENTER,Keys.TAB,par)
    elif(sheet.cell(i,6).value=="Working professional/Job seeker"):
        city.send_keys(Keys.TAB,3*Keys.UP,Keys.ENTER)
    elif(sheet.cell(i,6).value=="Job seeker"):
        city.send_keys(Keys.TAB,2*Keys.UP,Keys.ENTER)
    driver.find_element_by_xpath("//textarea[@label=\"Detailed Experience/ Pain points they mentioned\"]").send_keys("{}".format(sheet.cell(i,7).value))
    #driver.find_element_by_xpath("//textarea[@label=\"What did you share that would help them to address their pain points? (Your Experiences/Knowledge)\"]").send_keys("1768")
    #driver.find_element_by_xpath("//textarea[@label=\"Anything else you would like to add?\"]").send_keys("1768")
    driver.find_element_by_xpath("//label[@class=\"sc-iqzUVk kboBDW\"]").click()
    driver.find_element_by_xpath("//button[@title=\"Add Connection\"]").click()
    sleep(3)
    driver.find_element_by_xpath("//div[@class=\"styles_checkBoxImage__169Tr\"]").click()
    driver.find_element_by_xpath("//button[@id=\"56\"]").click()
    sleep(3)
#chrome.exe -remote-debugging-port=8080 --user-data-dir="C:\Users\anush\Download\Harsha\Class\Py\Bot\v2"
